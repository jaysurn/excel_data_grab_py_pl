from openpyxl import Workbook, load_workbook # Excel libs
# install lib 2.5.14 (the stable build)
import os # Access Directory libs
import re # Regex libs
def main():
	num_files = 1

	output_wb = Workbook()
	output_ws = output_wb.active
	data = []
	
	print( "Hello !" )
	direc = input( "Please type/paste in the full name of the directory you wish to submit\n" )
	output_file = os.getcwd() + "/output_wb.xlsx"
	output_wb.save( output_file )

	for files in os.listdir( direc ) :
		iterator = 1
		print ( files )
		full_path = direc + "/" + files
		wb = load_workbook( full_path, data_only = True )
		ws = wb[ "Event Information" ]
		
		# get line of data from input excel
		for row in ws.iter_rows ( min_row = 18,max_col = 26, max_row = 18, ) :
			for cell in row :
				good_cell = re.search( "[^245]", str( iterator ) )
				if good_cell:
					if ( iterator > 9 and str( cell.value ) == "0" ):
						temp_var = ""
					else:
						temp_var = cell.value
					data.append( temp_var )
		
				iterator += 1
		# output line of data to output excel
		for i in range (1,20 ) :
			output_cell_ref = output_ws.cell( row = num_files, column = i )
			
			# reformat time data
			if ( i == 1 ):
				temp = str( data[i-1] )
				temp = temp.strip(" 00:00:00")
				date = temp.split( "-" )
				data[i-1] = date[1] + "/" + date[2] + "/" + date[0]
			output_cell_ref.value = str( data[i-1] )
		
		# next file and clear data
		data = []
		num_files += 1
	
	output_wb.save( output_file )
		
main()